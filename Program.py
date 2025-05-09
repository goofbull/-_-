from PySide6.QtWidgets import * 
from untitled_ui import Ui_MainWindow
from export_to_excel import create_single_excel
from excel_to_csv_converter import convert
from prepare_csv import prepare_csv
from create_density_diagram import show_density
from rf_kmeans_pipeline import accuracy, precision, f1, recall, show_new_prediction
import txt_to_pdf
import openpyxl
import time
from contextlib import redirect_stdout
import os
import platform
import subprocess
import shutil


class MyApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
    
        self.ui.InputFilenameButton.clicked.connect(self.get_filename)
        self.ui.ShowDataButton.clicked.connect(self.load_data)
        self.ui.ShowResultButton.clicked.connect(self.show_result)
        self.ui.ShowDiagramsButton.clicked.connect(self.show_density_diagram)
        self.ui.ShowReportButton.clicked.connect(self.show_report)
        self.ui.SaveButton.clicked.connect(self.save_result)

    def get_filename(self):
        filename = self.ui.filename.text()
        print("Введённый файл:", filename)


    def load_data(self):
        with open('output.txt', 'w', encoding='utf-8') as f:
            with redirect_stdout(f):
                start_time = time.time()
                create_single_excel(self.ui.filename.text())

                path = "excel_files/" + self.ui.filename.text() + ".xlsx"
                workbook = openpyxl.load_workbook(path)
                sheet = workbook.active

                self.ui.tableWidget.setRowCount(sheet.max_row)
                self.ui.tableWidget.setColumnCount(sheet.max_column)

                list_values = list(sheet.values)
                self.ui.tableWidget.setHorizontalHeaderLabels(list_values[0])

                row_index = 0
                for value_tuple in list_values[1:]:
                    col_index = 0
                    for value in value_tuple:
                        self.ui.tableWidget.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                        col_index += 1
                    row_index += 1
                end_time = time.time()
                execution_time = end_time - start_time
                print("Загрузка данных...")
                print(f"Время выполнения операции: {execution_time:.4f} секунд")
                print("Excel файл готов")

    def show_result(self):
        with open('output.txt', 'a', encoding='utf-8') as f:
            with redirect_stdout(f):
                start_time = time.time()
                print("Конвертирование в csv файл...")
                convert(self.ui.filename.text())
                print("Подготовка csv файла для анализа...")
                prepare_csv(self.ui.filename.text())
                print("Вычисление результата...")
                prediction = show_new_prediction(self.ui.filename.text())
                self.ui.Result.setText(str(prediction))
                print("Предсказание: ", prediction)

                print("Метрики: ")
                print("Accuracy: ", accuracy)
                print("Precision: ", precision)
                print("Recall: ", recall)
                print("F1 Score: ", f1)

                self.ui.accuracy.setText(str(accuracy))
                self.ui.precision.setText(str(precision))
                self.ui.f1_score.setText(str(f1))
                self.ui.recall.setText(str(recall))
                end_time = time.time()
                execution_time = end_time - start_time
                print(f"Время выполнения операции: {execution_time:.4f} секунд")

    def save_result(self):
        source_pdf_path = "output.pdf"
        source_pdf_path_diag = "./images/fig1.pdf"

        output_dir = "saved_results"
        os.makedirs(output_dir, exist_ok=True)

        pdf_path = os.path.join(output_dir, "output.pdf")
        pdf_path_diag = os.path.join(output_dir, 'fig1.pdf')
        
        shutil.move(source_pdf_path, pdf_path)
        shutil.move(source_pdf_path_diag, pdf_path_diag)

    def show_density_diagram(self):
        with open('output.txt', 'a', encoding='utf-8') as f:
            with redirect_stdout(f):
                start_time = time.time()
                show_density(self.ui.filename.text())
                end_time = time.time()
                execution_time = end_time - start_time
                print("Загрузка графика...")
                print(f"Время выполнения операции: {execution_time:.4f} секунд")

    def show_report(self):
        txt_to_pdf.txt_to_pdf_fpdf("output.txt", "output.pdf")
        self.open_pdf()

    def open_pdf(self):
        pdf_path = "output.pdf"
        if platform.system() == "Windows":
            os.startfile(pdf_path)
        elif platform.system() == "Darwin":  # macOS
            subprocess.run(["open", pdf_path])
        else:  # Linux
            subprocess.run(["xdg-open", pdf_path])


if __name__ == "__main__":
    app = QApplication([])
    window = MyApp()
    window.show()
    app.exec()


